import pandas as pd
import os
import shutil
from datetime import datetime
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton, 
                            QTableView, QHeaderView, QMessageBox, QSpinBox, QLabel, QFileDialog, QLineEdit, QComboBox)
from PyQt5.QtCore import Qt, QAbstractTableModel, QVariant
import subprocess
import pprint
import numpy as np
from openpyxl import load_workbook

class PandasModel(QAbstractTableModel):
    def __init__(self, df, editable_col=None, excel_header_row=None):
        super().__init__()
        self._df = df.copy()
        self.editable_col = editable_col
        self.excel_header_row = excel_header_row

    def rowCount(self, parent=None):
        return self._df.shape[0]

    def columnCount(self, parent=None):
        return self._df.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return QVariant()
        # Для первой строки показываем excel_header_row, если есть
        if self.excel_header_row is not None and index.row() == 0 and role in (Qt.DisplayRole, Qt.EditRole):
            val = self.excel_header_row[index.column()]
            return str(val) if val is not None else ''
        value = self._df.iloc[index.row(), index.column()]
        if role == Qt.DisplayRole or role == Qt.EditRole:
            return str(value) if pd.notnull(value) else ''
        return QVariant()

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(self._df.columns[section])
            else:
                return str(section + 1)
        return QVariant()

    def flags(self, index):
        if not index.isValid():
            return Qt.ItemIsEnabled
        if self.editable_col is not None and self._df.columns[index.column()] == self.editable_col:
            return Qt.ItemFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        return Qt.ItemFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    def setData(self, index, value, role=Qt.EditRole):
        if self.editable_col is not None and self._df.columns[index.column()] == self.editable_col:
            try:
                val = int(value)
            except Exception:
                val = 0
            self._df.iloc[index.row(), index.column()] = val
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
            return True
        return False

    def get_dataframe(self):
        return self._df.copy()

def load_excel_with_merged_cells(file_path, sheet_name=None, header_row=0):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]
    # Разворачиваем объединённые ячейки
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = value
    # Считываем данные в список списков
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    # Преобразуем в DataFrame
    import pandas as pd
    df = pd.DataFrame(data)
    # Первая строка после header_row — это заголовки
    df.columns = df.iloc[header_row]
    df = df.drop(index=list(range(0, header_row+1)))
    df = df.reset_index(drop=True)
    # Сохраняем промежуточный результат для диагностики
    try:
        df.to_csv('temp_debug_output.csv', index=False, encoding='utf-8-sig')
        print('DEBUG CSV SAVED!')
    except Exception as e:
        print(f'Ошибка при сохранении debug csv: {e}')
    return df

class PriceListProcessor:
    def __init__(self, db):
        self.db = db
        self.price_list_data = None
        self.selected_items = {}
        self.column_mapping = {}

    def find_header_row(self, file_path, max_scan_rows=30):
        """Автоматически ищет строку с заголовками по ключевым словам."""
        # Ключевые слова для поиска
        keywords = [
            'артикул', 'название', 'описание', 'характерист', 'размер', 'цена', 'оптов', 'категор', 'группа', 'photo', 'фото', 'упаков', 'order', 'заказ', 'сумма'
        ]
        xl = pd.ExcelFile(file_path)
        for sheet in xl.sheet_names:
            df = xl.parse(sheet, header=None, nrows=max_scan_rows)
            for i, row in df.iterrows():
                row_str = ' '.join([str(cell).lower() for cell in row if pd.notnull(cell)])
                if sum(kw in row_str for kw in keywords) >= 2:  # если найдено хотя бы 2 ключевых слова
                    print(f"Заголовок найден на строке {i} листа '{sheet}': {row_str}")
                    return i, sheet
        raise ValueError("Не удалось найти строку с заголовками таблицы. Проверьте структуру файла.")

    def detect_columns(self, df):
        """Определение нужных колонок по их содержимому (улучшено: гибкий поиск, добавлен 'фото и описание')"""
        possible_names = {
            'name': [
                'название', 'товар', 'продукт', 'name', 'product', 'item', 'артикул', 'наименование', 'назв', 'описание', 'description', 'фото и описание'
            ],
            'category': ['категория', 'группа', 'category', 'group', 'тип', 'вид', 'раздел', 'подраздел']
        }
        print("\nДоступные колонки в файле:")
        for col in df.columns:
            print(f"- {col}")
        # Поиск колонки цены с учетом сложных вариантов
        price_col = None
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if 'цена' in col_lower and any(x in col_lower for x in ['руб', 'price', 'cost']):
                price_col = col
                break
        if not price_col:
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if 'цена' in col_lower:
                    price_col = col
                    break
        if price_col:
            self.column_mapping['price'] = price_col
            print(f"Найдено соответствие для price: {price_col}")
        # Поиск остальных колонок
        for col in df.columns:
            col_lower = str(col).strip().lower()
            for key, possible_values in possible_names.items():
                if key == 'price':
                    continue
                if any(val in col_lower for val in possible_values):
                    print(f"Найдено соответствие для {key}: {col}")
                    self.column_mapping[key] = col
                    break
        required_columns = ['name', 'price']
        missing_columns = [col for col in required_columns if col not in self.column_mapping]
        if missing_columns:
            print("\nНе найдены следующие колонки:")
            for col in missing_columns:
                print(f"- {col}")
            raise ValueError(f"Не удалось определить колонки: {', '.join(missing_columns)}")
        if 'category' not in self.column_mapping:
            print("\nВнимание: колонка 'category' не найдена. Все товары будут с категорией 'Без категории'.")
        return True

    def load_price_list(self, file_path):
        """Загрузка прайс-листа из Excel файла с улучшенной обработкой данных и автозаполнением объединённых ячеек."""
        try:
            print(f"\nЗагрузка файла: {file_path}")
            
            # 1. Определяем строку заголовка
            header_row, sheet = self.find_header_row(file_path)
            print(f"Используем строку {header_row} как заголовок на листе '{sheet}'")
            
            # 2. Загружаем данные с учетом объединенных ячеек
            df = load_excel_with_merged_cells(file_path, sheet_name=sheet, header_row=header_row)
            
            # 3. Очистка данных
            df = df.dropna(how='all').dropna(axis=1, how='all')
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace(r'^\s*$', np.nan, regex=True)
            
            # 4. Определяем важные колонки
            cols = [str(c).lower() for c in df.columns]
            name_col = next((c for c in df.columns if any(x in str(c).lower() for x in 
                ['назв', 'name', 'товар', 'product', 'item', 'артикул', 'наименование'])), None)
            price_col = next((c for c in df.columns if any(x in str(c).lower() for x in 
                ['цена', 'price', 'cost', 'стоимость', 'руб', '₽'])), None)
            art_col = next((c for c in df.columns if any(x in str(c).lower() for x in 
                ['артикул', 'sku', 'код', 'code', 'номер'])), None)

            # 4.1. Заполняем пустые значения в важных колонках методом ffill
            for col in [name_col, art_col]:
                if col and col in df.columns:
                    df[col] = df[col].fillna(method='ffill')
            
            # 5. Проверяем наличие обязательных колонок
            if not name_col and not art_col:
                raise ValueError("Не найдены колонки с названием или артикулом товара")
            if not price_col:
                raise ValueError("Не найдена колонка с ценой")
            
            # 6. Обработка цен
            if price_col:
                df[price_col] = df[price_col].astype(str).str.replace(r'[^\d.,]', '', regex=True)
                df[price_col] = df[price_col].str.replace(',', '.')
                df[price_col] = pd.to_numeric(df[price_col], errors='coerce')
                df = df[df[price_col] > 0]
            
            # 7. Обработка названий и артикулов
            if name_col:
                df = df.drop_duplicates(subset=[name_col])
                df = df[df[name_col].notna() & (df[name_col] != '')]
            if art_col:
                df = df.drop_duplicates(subset=[art_col])
                df = df[df[art_col].notna() & (df[art_col] != '')]
            
            # 8. Сохраняем промежуточный результат для диагностики
            df.to_csv('temp_debug_output.csv', index=False, encoding='utf-8-sig')
            print("\nОбработанные данные:")
            print(f"Всего строк: {len(df)}")
            print(f"Колонки: {', '.join(df.columns)}")
            print("\nПервые несколько строк:")
            print(df.head())
            
            # 9. Определяем и сохраняем маппинг колонок
            self.column_mapping = {
                'name': name_col,
                'price': price_col,
                'article': art_col
            }
            
            # 10. Переименовываем колонки для удобства
            rename_dict = {
                name_col: 'Название',
                price_col: 'Цена',
            }
            if art_col:
                rename_dict[art_col] = 'Артикул'
            df = df.rename(columns=rename_dict)
            
            # 11. Добавляем категорию, если её нет
            if 'Категория' not in df.columns:
                df['Категория'] = 'Без категории'
            
            self.price_list_data = df
            return True
            
        except Exception as e:
            print(f"Ошибка при загрузке прайс-листа: {e}")
            return False

    def find_order_column(self, df):
        """Автоматически ищет колонку для заказа по ключевым словам."""
        order_keywords = ['заказ', 'order', 'укажите', 'количество', 'qty', 'шт', 'штук']
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in order_keywords):
                return col
        raise ValueError("Не удалось найти колонку для заказа. Добавьте колонку 'Заказ' или аналогичную в прайс-лист.")

    def create_order_file(self, file_path, selected_items):
        """Создаёт копию исходного прайс-листа и записывает в ней количество заказа для выбранных товаров."""
        try:
            # Копируем исходный файл
            base, ext = os.path.splitext(file_path)
            output_path = base + '_order' + ext
            shutil.copy(file_path, output_path)
            print(f"Создана копия файла: {output_path}")

            # Загружаем копию
            xl = pd.ExcelFile(output_path)
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                try:
                    order_col = self.find_order_column(df)
                except Exception:
                    continue
                # Определяем уникальный идентификатор (артикул или название)
                id_col = None
                for col in df.columns:
                    if any(x in str(col).lower() for x in ['артикул', 'sku', 'код', 'id', 'название', 'name']):
                        id_col = col
                        break
                if not id_col:
                    continue
                # Записываем количество заказа
                for idx, row in df.iterrows():
                    item_id = str(row[id_col]).strip()
                    if item_id in selected_items:
                        df.at[idx, order_col] = selected_items[item_id]
                # Сохраняем обратно на этот же лист
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet, index=False)
            return True
        except Exception as e:
            print(f"Ошибка при создании файла заказа: {e}")
            return False

    def add_to_pending_orders(self, selected_items, supplier_id=None):
        """Добавление выбранных товаров в ожидающие заказы (master-detail)"""
        try:
            # 1. Добавляем заказ (шапку)
            self.db.cursor.execute("""
                INSERT INTO pending_orders (supplier, order_date, status)
                VALUES (%s, %s, %s) RETURNING id
            """, (supplier_id, datetime.now(), 'Ожидает поступления'))
            order_id = self.db.cursor.fetchone()[0]
            # 2. Добавляем товары (позиции)
            for item_name, quantity in selected_items.items():
                item_data = self.price_list_data[self.price_list_data['Название'] == item_name].iloc[0]
                # Проверка на пустые значения
                if (not item_name or pd.isnull(item_name) or str(item_name).strip() == '' or
                    pd.isnull(item_data['Цена']) or str(item_data['Цена']).strip() == '' or
                    pd.isnull(item_data['Категория']) or str(item_data['Категория']).strip() == ''):
                    print(f"[SKIP] Пустое значение: name={item_name}, price={item_data['Цена']}, category={item_data['Категория']}")
                    continue
                self.db.cursor.execute("""
                    INSERT INTO pending_order_items (order_id, name, price, quantity, category)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    order_id,
                    item_name,
                    float(item_data['Цена']),
                    quantity,
                    item_data['Категория']
                ))
            self.db.connection.commit()
            return True
        except Exception as e:
            print(f"Ошибка при добавлении в ожидающие заказы: {e}")
            self.db.connection.rollback()
            return False

class ColumnMappingDialog(QDialog):
    def __init__(self, df, parent=None, excel_file=None):
        super().__init__(parent)
        self._original_df = df.reset_index(drop=True)
        self.df = self._original_df.copy()
        self.excel_file = excel_file
        self.excel_header_row = None  # Сохраняем значения выбранной строки Excel
        self.setWindowTitle("Сопоставление колонок")
        self.showMaximized()
        layout = QVBoxLayout(self)
        self.header_row_spin = QSpinBox()
        self.header_row_spin.setMinimum(0)
        self.header_row_spin.setMaximum(0)
        self.header_row_spin.setPrefix("Строка с названиями: ")
        self.header_row_spin.valueChanged.connect(self.on_header_row_changed)
        layout.addWidget(self.header_row_spin)
        self.mapping_options = [
            "Пропустить", "Название", "Артикул", "Цена", "Количество", "Категория"
        ]
        self.combos = []
        self.combo_layout = QHBoxLayout()
        layout.addLayout(self.combo_layout)
        self.table_view = QTableView()
        self.model = None  # инициализация позже
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table_view)

        # Контекстное меню для удаления столбца
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_table_context_menu)
        self.table_view.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.horizontalHeader().customContextMenuRequested.connect(self.show_header_context_menu)

        # Кнопка заполнения пустых ячеек сверху
        self.ffill_btn = QPushButton("Заполнить пустые ячейки сверху")
        self.ffill_btn.setStyleSheet("background-color: #43e97b; color: #23243a; font-weight: bold; padding: 8px 18px; border-radius: 8px;")
        self.ffill_btn.clicked.connect(self.fill_empty_from_above)
        layout.addWidget(self.ffill_btn)

        # Кнопка удаления выбранных строк
        self.delete_rows_btn = QPushButton("Удалить выбранные строки")
        self.delete_rows_btn.setStyleSheet("background-color: #ff5555; color: white; font-weight: bold; padding: 8px 18px; border-radius: 8px;")
        self.delete_rows_btn.clicked.connect(self.delete_selected_rows)
        layout.addWidget(self.delete_rows_btn)

        btns = QHBoxLayout()
        self.ok_btn = QPushButton("Подтвердить")
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.ok_btn)
        btns.addWidget(self.cancel_btn)
        layout.addLayout(btns)
        if self.excel_file is not None:
            import pandas as pd
            xl = pd.ExcelFile(self.excel_file)
            sheet = xl.sheet_names[0]
            df_raw = xl.parse(sheet, header=None)
            self._df_raw = df_raw
            self.header_row_spin.setMaximum(len(df_raw)-1)
            for i, row in df_raw.iterrows():
                if sum(str(cell).strip() not in ('', 'nan', 'NaN') for cell in row) >= 2:
                    self.header_row_spin.setValue(i)
                    break
            self.on_header_row_changed()
        else:
            self.update_combos_and_headers(self.df.columns)
            self.model = PandasModel(self.df)
            self.table_view.setModel(self.model)

    def on_header_row_changed(self):
        idx = self.header_row_spin.value()
        row = list(self._df_raw.iloc[idx])
        # Проверяем, есть ли хотя бы 2 непустых значения
        non_empty = [str(cell).strip() for cell in row if str(cell).strip() and str(cell).strip().lower() != 'nan']
        if len(non_empty) >= 2:
            headers = []
            for i, cell in enumerate(row):
                val = str(cell).strip()
                if not val or val.lower() == 'nan':
                    headers.append(f'Unnamed: {i}')
                else:
                    headers.append(val)
            if len(headers) == len(self.df.columns):
                self.df.columns = headers
        # Обновляем комбобоксы и таблицу
        self.update_combos_and_headers(self.df.columns)
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

    def update_combos_and_headers(self, headers):
        # Очищаем старые комбобоксы из combo_layout
        while self.combo_layout.count():
            item = self.combo_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)
        self.combos = []
        prev_choices = [combo.currentText() for combo in self.combos] if self.combos else []
        # Используем только реальные имена столбцов DataFrame
        for idx, col in enumerate(self.df.columns):
            combo = QComboBox()
            combo.addItems(self.mapping_options)
            # Восстанавливаем выбор, если был
            if idx < len(prev_choices):
                prev = prev_choices[idx]
                if prev in self.mapping_options:
                    combo.setCurrentText(prev)
            combo.currentIndexChanged.connect(self.on_mapping_changed)
            self.combo_layout.addWidget(combo)
            self.combos.append(combo)
        # После обновления модели подгоняем высоту таблицы под количество строк
        self.table_view.resizeRowsToContents()
        max_visible_rows = 20
        row_height = self.table_view.verticalHeader().defaultSectionSize()
        num_rows = min(self.df.shape[0], max_visible_rows)
        table_height = row_height * (num_rows + 1)  # +1 для заголовка
        self.table_view.setMaximumHeight(table_height)
        from PyQt5.QtWidgets import QSizePolicy
        self.table_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)

    def on_mapping_changed(self):
        mapping = [combo.currentText() for combo in self.combos]
        required_fields = ["Название", "Цена"]
        missing_required = [field for field in required_fields if field not in mapping]
        if missing_required:
            QMessageBox.warning(self, "Ошибка", f"Не выбраны обязательные поля: {', '.join(missing_required)}")
            return
        else:
            if "Название" in mapping:
                name_col_idx = mapping.index("Название")
                if name_col_idx < len(self.df.columns):
                    orig_col_name = self.df.columns[name_col_idx]
                    mask = self.df[orig_col_name].apply(lambda x: str(x).strip() not in ('', 'nan', 'NaN'))
                    if len(mask) == len(self.df):
                        filtered = self.df.loc[mask]
                        self.df = filtered.reset_index(drop=True)
            if "Количество" in mapping:
                qty_col_idx = mapping.index("Количество")
                if qty_col_idx < len(self.df.columns):
                    qty_col_name = self.df.columns[qty_col_idx]
                    def is_valid_qty(x):
                        try:
                            return str(x).strip() not in ('', 'nan', 'NaN') and float(x) != 0
                        except Exception:
                            return False
                    mask = self.df[qty_col_name].apply(is_valid_qty)
                    if len(mask) == len(self.df):
                        filtered = self.df.loc[mask]
                        self.df = filtered.reset_index(drop=True)
        # Не меняем заголовки DataFrame!
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

    def get_mapping(self):
        return [combo.currentText() for combo in self.combos]

    def delete_selected_rows(self):
        selected = self.table_view.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "Удаление строк", "Выделите одну или несколько строк для удаления.")
            return
        idxs = sorted([index.row() for index in selected], reverse=True)
        self.df = self.df.drop(self.df.index[idxs]).reset_index(drop=True)
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

    def fill_empty_from_above(self):
        self.df = self.df.ffill().reset_index(drop=True)
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

    def show_header_context_menu(self, pos):
        from PyQt5.QtWidgets import QMenu
        header = self.table_view.horizontalHeader()
        logical_index = header.logicalIndexAt(pos)
        if logical_index < 0:
            return
        col_name = self.df.columns[logical_index]
        menu = QMenu(self)
        delete_action = menu.addAction(f"Удалить столбец '{col_name}' (#{logical_index+1})")
        # Новое действие: удалить все выделенные столбцы
        selected_cols = sorted(set([header.logicalIndex(i.column()) for i in header.selectionModel().selectedIndexes()]))
        if len(selected_cols) > 1:
            delete_multi_action = menu.addAction(f"Удалить выделенные столбцы ({len(selected_cols)})")
        else:
            delete_multi_action = None
        action = menu.exec_(header.mapToGlobal(pos))
        if action == delete_action:
            # Удаляем только выбранный столбец по индексу
            self.df = self.df.iloc[:, [i for i in range(self.df.shape[1]) if i != logical_index]]
            self.model = PandasModel(self.df)
            self.table_view.setModel(self.model)
            self.update_combos_and_headers(self.df.columns)
        elif delete_multi_action and action == delete_multi_action:
            # Удаляем все выделенные столбцы
            keep = [i for i in range(self.df.shape[1]) if i not in selected_cols]
            self.df = self.df.iloc[:, keep]
            self.model = PandasModel(self.df)
            self.table_view.setModel(self.model)
            self.update_combos_and_headers(self.df.columns)

    def show_table_context_menu(self, pos):
        from PyQt5.QtWidgets import QMenu, QAction, QMessageBox
        index = self.table_view.indexAt(pos)
        if not index.isValid():
            return  # Клик вне строки/ячейки
        selected = self.table_view.selectionModel().selectedRows()
        if not selected:
            return  # Нет выделения — ничего не делать
        menu = QMenu(self)
        delete_action = QAction("Удалить выбранные строки", self)
        menu.addAction(delete_action)
        action = menu.exec_(self.table_view.viewport().mapToGlobal(pos))
        if action == delete_action:
            self.delete_selected_rows()

class PriceListDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.processor = PriceListProcessor(db)
        self.setWindowTitle("Обработка прайс-листа")
        self.original_file = None
        self.edit_file = None
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.load_button = QPushButton("Выбрать прайс-лист и открыть для редактирования")
        self.load_button.clicked.connect(self.load_and_open_file)
        layout.addWidget(self.load_button)
        self.confirm_button = QPushButton("Подтвердить изменения и обработать заказ")
        self.confirm_button.clicked.connect(self.confirm_changes)
        self.confirm_button.setEnabled(False)
        layout.addWidget(self.confirm_button)
        self.changes_label = QLabel()
        self.changes_label.setWordWrap(True)
        layout.addWidget(self.changes_label)
        self.setLayout(layout)

    def load_and_open_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите прайс-лист",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.original_file = file_path
            base, ext = os.path.splitext(file_path)
            self.edit_file = base + '_edit' + ext
            shutil.copy(file_path, self.edit_file)
            # Открываем копию в Excel
            try:
                os.startfile(self.edit_file)
            except Exception:
                try:
                    subprocess.Popen(['start', self.edit_file], shell=True)
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось открыть файл: {e}")
                    return
            self.confirm_button.setEnabled(True)
            self.changes_label.setText("Внесите изменения в открытом файле, затем сохраните его и нажмите 'Подтвердить изменения и обработать заказ'.")

    def confirm_changes(self):
        # Сравниваем исходный и изменённый файлы
        if not self.original_file or not self.edit_file:
            QMessageBox.warning(self, "Внимание", "Сначала выберите и откройте прайс-лист!")
            return
        try:
            orig = pd.read_excel(self.original_file)
            edited = pd.read_excel(self.edit_file)
            # Сохраняем изменённый DataFrame для диагностики
            try:
                edited.to_csv('debug_edited_price_list.csv', index=False, encoding='utf-8-sig')
                print('DEBUG CSV (edited price list) SAVED!')
            except Exception as e:
                print(f'Ошибка при сохранении debug csv: {e}')
            min_len = min(len(orig), len(edited))
            orig = orig.iloc[:min_len]
            edited = edited.iloc[:min_len]
            changed_rows = []
            for idx in range(min_len):
                row_changed = False
                for col in orig.columns:
                    if col in edited.columns:
                        val_orig = orig.iloc[idx][col]
                        val_edit = edited.iloc[idx][col]
                        if pd.isnull(val_orig) and pd.isnull(val_edit):
                            continue
                        if str(val_orig) != str(val_edit):
                            row_changed = True
                            break
                if row_changed:
                    changed_rows.append(edited.iloc[idx])
            if not changed_rows:
                self.changes_label.setText("Изменений не обнаружено. Проверьте, что вы сохранили файл после редактирования.")
                return
            changed_df = pd.DataFrame(changed_rows)
            # Открываем диалог сопоставления колонок
            mapping_dialog = ColumnMappingDialog(changed_df, self, self.original_file)
            if mapping_dialog.exec_() == QDialog.Accepted:
                mapping = mapping_dialog.get_mapping()
                col_map = dict(zip(changed_df.columns, mapping))
                self.col_map = col_map  # сохраняем для create_order
                self.mapping = mapping
                QMessageBox.information(self, "Сопоставление завершено", f"Сопоставление колонок: {col_map}")
                # После сопоставления колонок запускаем оформление заказа
                self.model = mapping_dialog.model  # чтобы create_order работал с нужной таблицей
                self.create_order()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сравнить файлы: {e}")

    def filter_table(self, text):
        if self.df is not None:
            filtered = self.df[self.df.apply(lambda row: row.astype(str).str.contains(text, case=False).any(), axis=1)]
            self.model = PandasModel(filtered, editable_col='Количество')
            self.table_view.setModel(self.model)

    def create_order(self):
        # Используем только сопоставленные столбцы (не "Пропустить")
        df = self.model.get_dataframe()
        col_map = getattr(self, 'col_map', None)
        if not col_map:
            QMessageBox.critical(self, "Ошибка", "Не удалось получить сопоставление колонок!")
            return
        # Определяем реальные имена столбцов для Название, Цена, Количество, Артикул и т.д.
        name_col = None
        price_col = None
        qty_col = None
        art_col = None
        for col, mapped in col_map.items():
            if mapped == "Название":
                name_col = col
            elif mapped == "Цена":
                price_col = col
            elif mapped == "Количество":
                qty_col = col
            elif mapped == "Артикул":
                art_col = col
        if not name_col or not price_col or not qty_col:
            QMessageBox.critical(self, "Ошибка", "Не выбраны все обязательные поля (Название, Цена, Количество)!")
            return
        # Отладочный вывод: имена столбцов и сопоставление
        debug_msg = f"Столбцы: {list(df.columns)}\n"
        debug_msg += f"name_col: {name_col}\nprice_col: {price_col}\nqty_col: {qty_col}\nart_col: {art_col}\n"
        if not df.empty:
            debug_msg += f"Первая строка: {df.iloc[0].to_dict()}\n"
        else:
            debug_msg += "Таблица пуста\n"
        QMessageBox.information(self, "Отладка", debug_msg)
        selected_items = {}
        for idx, row in df.iterrows():
            try:
                qty_val = row[qty_col]
                if isinstance(qty_val, str):
                    qty_val = qty_val.replace(',', '.').strip()
                qty = int(float(qty_val))
            except Exception:
                qty = 0
            if qty > 0:
                # Ключ — артикул, если есть, иначе название
                if art_col:
                    key = str(row[art_col])
                else:
                    key = str(row[name_col])
                selected_items[key] = {
                    'qty': qty,
                    'name': str(row[name_col]),
                    'price': float(row[price_col]) if pd.notnull(row[price_col]) else 0.0
                }
        # Показываем отладку даже если selected_items пустой
        QMessageBox.information(self, "Отладка: товары", f"selected_items:\n{pprint.pformat(selected_items)}")
        if not selected_items:
            QMessageBox.warning(self, "Предупреждение", "Не выбрано ни одного товара")
            return
        # Показываем диалог выбора поставщика
        from app_code.warehouse_page import SupplierSelectDialog
        supplier_dialog = SupplierSelectDialog(self.db, self)
        if supplier_dialog.exec_() != QDialog.Accepted:
            return
        supplier_id, supplier_name = supplier_dialog.get_selected_supplier()
        # Сразу добавляем заказ в базу, без сохранения файла
        items_for_order = {k: v['qty'] for k, v in selected_items.items()}
        # Переименовываем столбцы в стандартные имена для дальнейшей работы
        rename_dict = {}
        if name_col:
            rename_dict[name_col] = 'Название'
        if price_col:
            rename_dict[price_col] = 'Цена'
        if 'Категория' not in df.columns and any(mapped == "Категория" for mapped in col_map.values()):
            for col, mapped in col_map.items():
                if mapped == "Категория":
                    rename_dict[col] = 'Категория'
        df = df.rename(columns=rename_dict)
        if 'Категория' not in df.columns:
            df['Категория'] = 'Без категории'
        self.processor.price_list_data = df
        if self.processor.add_to_pending_orders(items_for_order, supplier_id):
            QMessageBox.information(
                self,
                "Успех",
                f"Заказ успешно создан для поставщика: {supplier_name} и добавлен в ожидающие"
            )
            self.accept()
        else:
            QMessageBox.critical(
                self,
                "Ошибка",
                "Не удалось добавить товары в ожидающие заказы"
            ) 