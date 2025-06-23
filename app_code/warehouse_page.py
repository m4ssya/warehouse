from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                            QPushButton, QFrame, QScrollArea, QTableWidget, 
                            QTableWidgetItem, QHeaderView, QMessageBox, QDialog,
                            QLineEdit, QFormLayout, QSpinBox, QComboBox, QCheckBox,
                            QMenu, QAction, QListWidget, QListWidgetItem, QInputDialog,
                            QSizePolicy)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QColor, QFont
import datetime
from app_code.warehouse_automation import WarehouseAutomation
from app_code.price_list_processor import PriceListDialog, ColumnMappingDialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

class AddProductDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Добавление товара")
        self.setModal(True)
        self.resize(480, 420)
        self.setStyleSheet("""
            QDialog {
                background-color: #23243a;
                border-radius: 18px;
                border: 2px solid #43e97b;
            }
            QLineEdit, QComboBox, QSpinBox {
                background: #292a3a;
                color: #f3f3f3;
                border-radius: 10px;
                border: 1.5px solid #43e97b;
                font-size: 18px;
                padding: 8px 14px;
                margin-bottom: 8px;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
                border: 2px solid #43e97b;
                background: #23243a;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #43e97b, stop:1 #2fd072);
                color: #23243a;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
                padding: 10px 0;
                margin: 8px 0 0 0;
            }
            QPushButton:hover {
                background: #43e97b;
                color: #23243a;
            }
            QLabel {
                font-size: 17px;
            }
            QPushButton, QLineEdit, QComboBox, QSpinBox, QTableWidget, QTableWidget::item {
                outline: none;
            }
            QPushButton:focus, QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QTableWidget:focus {
                border: 2px solid #3fa996;
                box-shadow: 0 0 0 2px #3fa99644;
            }
        """)
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout(self)
        
        # Поля ввода
        self.name_input = QLineEdit()
        self.purchase_price_input = QLineEdit()
        self.purchase_price_input.setPlaceholderText("Закупочная цена")
        self.retail_price_input = QLineEdit()
        self.retail_price_input.setPlaceholderText("Розничная цена")
        self.price_input = QLineEdit()
        self.price_input.setPlaceholderText("Цена (устар.)")
        self.quantity_input = QSpinBox()
        self.quantity_input.setMinimum(0)
        self.quantity_input.setMaximum(999999)
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("Введите или отсканируйте штрихкод")
        
        # Выбор категории
        self.category_combo = QComboBox()
        self.load_categories()
        
        # Добавляем поля в форму
        layout.addRow("Название:", self.name_input)
        layout.addRow("Закупочная цена:", self.purchase_price_input)
        layout.addRow("Розничная цена:", self.retail_price_input)
        layout.addRow("Цена (устар.)", self.price_input)
        layout.addRow("Количество:", self.quantity_input)
        layout.addRow("Штрихкод:", self.barcode_input)
        layout.addRow("Категория:", self.category_combo)
        
        # Кнопки
        buttons_layout = QHBoxLayout()
        self.save_button = QPushButton("Сохранить")
        self.cancel_button = QPushButton("Отмена")
        
        self.save_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        
        buttons_layout.addWidget(self.save_button)
        buttons_layout.addWidget(self.cancel_button)
        layout.addRow(buttons_layout)

    def load_categories(self):
        self.db.cursor.execute("SELECT name FROM categories")
        categories = self.db.cursor.fetchall()
        self.category_combo.addItems([cat[0] for cat in categories])

    def get_product_data(self):
        return {
            'name': self.name_input.text(),
            'purchase_price': self.purchase_price_input.text(),
            'retail_price': self.retail_price_input.text(),
            'price': self.price_input.text(),
            'quantity': str(self.quantity_input.value()),
            'barcode': self.barcode_input.text(),
            'category': self.category_combo.currentText()
        }

class AutomationSettingsDialog(QDialog):
    def __init__(self, automation, parent=None):
        super().__init__(parent)
        self.automation = automation
        self.setWindowTitle("Настройки автоматизации")
        self.setModal(True)
        self.resize(480, 320)
        self.setStyleSheet("""
            QDialog {
                background-color: #23243a;
                border-radius: 18px;
                border: 2px solid #43e97b;
            }
            QSpinBox, QCheckBox {
                font-size: 18px;
            }
            QSpinBox {
                background: #292a3a;
                color: #f3f3f3;
                border-radius: 10px;
                border: 1.5px solid #43e97b;
                padding: 8px 14px;
                margin-bottom: 8px;
            }
            QSpinBox:focus {
                border: 2px solid #43e97b;
                background: #23243a;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #43e97b, stop:1 #2fd072);
                color: #23243a;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
                padding: 10px 0;
                margin: 8px 0 0 0;
            }
            QPushButton:hover {
                background: #43e97b;
                color: #23243a;
            }
            QLabel {
                font-size: 17px;
            }
        """)
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout(self)
        
        # Автоматическое формирование заказов
        self.auto_order_check = QCheckBox()
        self.auto_order_check.setChecked(self.automation.settings['auto_order'])
        layout.addRow("Автоматическое формирование заказов:", self.auto_order_check)
        
        # Порог для заказа
        self.order_threshold = QSpinBox()
        self.order_threshold.setRange(1, 100)
        self.order_threshold.setValue(int(self.automation.settings['order_threshold'] * 100))
        layout.addRow("Порог для заказа (% от минимального):", self.order_threshold)
        
        # Интервал проверки
        self.check_interval = QSpinBox()
        self.check_interval.setRange(1, 60)
        self.check_interval.setValue(self.automation.settings['check_interval'] // 60)
        layout.addRow("Интервал проверки (минуты):", self.check_interval)
        
        # Уведомления о низком остатке
        self.notify_check = QCheckBox()
        self.notify_check.setChecked(self.automation.settings['notify_on_low'])
        layout.addRow("Уведомлять о низком остатке:", self.notify_check)
        
        # Кнопки
        buttons_layout = QHBoxLayout()
        self.save_button = QPushButton("Сохранить")
        self.cancel_button = QPushButton("Отмена")
        
        self.save_button.clicked.connect(self.save_settings)
        self.cancel_button.clicked.connect(self.reject)
        
        buttons_layout.addWidget(self.save_button)
        buttons_layout.addWidget(self.cancel_button)
        layout.addRow(buttons_layout)

    def save_settings(self):
        new_settings = {
            'auto_order': self.auto_order_check.isChecked(),
            'order_threshold': self.order_threshold.value() / 100,
            'check_interval': self.check_interval.value() * 60,
            'notify_on_low': self.notify_check.isChecked()
        }
        self.automation.update_settings(new_settings)
        self.accept()

class AddSupplierDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить поставщика")
        self.setModal(True)
        self.resize(480, 320)
        self.setStyleSheet("""
            QDialog {
                background-color: #23243a;
                border-radius: 18px;
                border: 2px solid #43e97b;
            }
            QLineEdit {
                background: #292a3a;
                color: #f3f3f3;
                border-radius: 10px;
                border: 1.5px solid #43e97b;
                font-size: 18px;
                padding: 8px 14px;
                margin-bottom: 8px;
            }
            QLineEdit:focus {
                border: 2px solid #43e97b;
                background: #23243a;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #43e97b, stop:1 #2fd072);
                color: #23243a;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
                padding: 10px 0;
                margin: 8px 0 0 0;
            }
            QPushButton:hover {
                background: #43e97b;
                color: #23243a;
            }
            QLabel {
                font-size: 17px;
            }
        """)
        self.setup_ui()

    def setup_ui(self):
        layout = QFormLayout(self)
        self.name_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.email_input = QLineEdit()
        self.comment_input = QLineEdit()
        layout.addRow("Название:", self.name_input)
        layout.addRow("Телефон:", self.phone_input)
        layout.addRow("Email:", self.email_input)
        layout.addRow("Комментарий:", self.comment_input)
        buttons_layout = QHBoxLayout()
        self.save_button = QPushButton("Сохранить")
        self.cancel_button = QPushButton("Отмена")
        self.save_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.save_button)
        buttons_layout.addWidget(self.cancel_button)
        layout.addRow(buttons_layout)

    def get_supplier_data(self):
        return {
            'name': self.name_input.text(),
            'phone': self.phone_input.text(),
            'email': self.email_input.text(),
            'comment': self.comment_input.text()
        }

class SupplierSelectDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Выбор поставщика")
        self.setModal(True)
        self.resize(420, 220)
        self.setStyleSheet("""
            QDialog {
                background-color: #23243a;
                border-radius: 18px;
                border: 2px solid #43e97b;
            }
            QComboBox {
                background: #292a3a;
                color: #f3f3f3;
                border-radius: 10px;
                border: 1.5px solid #43e97b;
                font-size: 18px;
                padding: 8px 14px;
                margin-bottom: 8px;
            }
            QComboBox:focus {
                border: 2px solid #43e97b;
                background: #23243a;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #43e97b, stop:1 #2fd072);
                color: #23243a;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
                padding: 10px 0;
                margin: 8px 0 0 0;
            }
            QPushButton:hover {
                background: #43e97b;
                color: #23243a;
            }
            QLabel {
                font-size: 17px;
            }
        """)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.supplier_combo = QComboBox()
        self.load_suppliers()
        layout.addWidget(QLabel("Выберите поставщика:"))
        layout.addWidget(self.supplier_combo)
        btns = QHBoxLayout()
        ok_btn = QPushButton("ОК")
        cancel_btn = QPushButton("Отмена")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)

    def load_suppliers(self):
        self.db.cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
        self.suppliers = self.db.cursor.fetchall()
        for sid, name in self.suppliers:
            self.supplier_combo.addItem(name, sid)

    def get_selected_supplier(self):
        idx = self.supplier_combo.currentIndex()
        if idx >= 0:
            return self.suppliers[idx]
        return None

class WarehousePage(QWidget):
    def __init__(self, db, username, parent=None):
        super().__init__(parent)
        self.db = db
        self.username = username
        self.setup_ui()
        self.setup_automation()
        self.load_products()

    def setup_automation(self):
        self.automation = WarehouseAutomation(self.db)
        self.automation.low_stock_alert.connect(self.on_low_stock)
        self.automation.order_needed.connect(self.on_order_needed)
        self.automation.stock_updated.connect(self.load_products)

    def setup_ui(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #23243a;
                color: #f3f3f3;
                font-family: 'Segoe UI', 'Arial', sans-serif;
                font-size: 17px;
            }
            QFrame, QGroupBox {
                background: #23243a;
                border-radius: 18px;
                border: 1.5px solid #3fa996;
                box-shadow: 0 4px 32px 0 rgba(63,169,150,0.10);
            }
            QTableWidget {
                background-color: #23243a;
                color: #f3f3f3;
                border-radius: 18px;
                border: 2px solid #3fa996;
                gridline-color: #34344a;
                selection-background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3fa996, stop:1 #2d7a6a);
                selection-color: #23243a;
                alternate-background-color: #292a3a;
                box-shadow: 0 4px 24px 0 rgba(63,169,150,0.08);
                font-size: 17px;
            }
            QTableWidget::item {
                padding: 14px 12px;
                border-radius: 12px;
                font-size: 17px;
                transition: background 0.2s;
            }
            QTableWidget::item:selected {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3fa996, stop:1 #2d7a6a);
                color: #23243a;
                font-weight: bold;
            }
            QTableWidget::item:hover {
                background-color: #3c3f56;
                transition: background 0.2s;
            }
            QHeaderView::section {
                background-color: #23243a;
                color: #3fa996;
                font-weight: bold;
                font-size: 20px;
                border: none;
                border-bottom: 3px solid #3fa996;
                padding: 16px 0 16px 0;
                border-radius: 0px;
                letter-spacing: 1px;
            }
            QTableCornerButton::section {
                background: #23243a;
                border: none;
            }
            QCheckBox {
                spacing: 8px;
                font-size: 17px;
            }
            QCheckBox::indicator {
                width: 24px;
                height: 24px;
                border-radius: 8px;
                border: 2px solid #43e97b;
                background: #23243a;
            }
            QCheckBox::indicator:checked {
                background: #43e97b;
                border: 2px solid #2fd072;
            }
            QComboBox, QLineEdit, QSpinBox {
                background-color: #292a3a;
                color: #f3f3f3;
                border: 2px solid #43e97b;
                border-radius: 12px;
                padding: 10px 16px;
                font-size: 17px;
                margin-right: 8px;
            }
            QComboBox:focus, QLineEdit:focus, QSpinBox:focus {
                border: 2.5px solid #2fd072;
                background-color: #23243a;
            }
            QComboBox QAbstractItemView {
                background: #23243a;
                color: #43e97b;
                selection-background-color: #43e97b;
                selection-color: #23243a;
                border-radius: 12px;
                font-size: 17px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3fa996, stop:1 #2d7a6a);
                color: #f3f3f3;
                font-weight: 700;
                border-radius: 10px;
                font-size: 15px;
                padding: 4px 10px;
                min-width: 120px;
                min-height: 28px;
                margin: 0;
                box-shadow: 0 2px 8px 0 rgba(63,169,150,0.10);
                transition: background 0.2s;
            }
            QPushButton:hover {
                background: #4fd1b3;
                color: #23243a;
                box-shadow: 0 4px 16px 0 rgba(63,169,150,0.18);
            }
            QLabel[role='title'] {
                font-size: 32px;
                font-weight: bold;
                color: #3fa996;
                margin-bottom: 18px;
                letter-spacing: 1.5px;
            }
            QToolTip {
                background-color: #23243a;
                color: #3fa996;
                border: 2px solid #3fa996;
                border-radius: 12px;
                font-size: 16px;
                padding: 10px;
            }
            QScrollBar:vertical, QScrollBar:horizontal {
                background: #23243a;
                border: none;
                border-radius: 8px;
                width: 14px;
                margin: 4px 0 4px 0;
            }
            QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3fa996, stop:1 #2d7a6a);
                min-height: 36px;
                border-radius: 8px;
                border: none;
                transition: background 0.2s;
            }
            QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover {
                background: #4fd1b3;
            }
            QScrollBar::add-line, QScrollBar::sub-line {
                background: none;
                border: none;
                height: 0px;
            }
            QScrollBar::add-page, QScrollBar::sub-page {
                background: none;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(24)

        # --- Горизонтальный контейнер для меню и заголовка ---
        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(12)
        # Если есть кнопка меню (например, self.menu_button), добавляем её слева
        if hasattr(self.parent(), 'menu_button'):
            top_row.addWidget(self.parent().menu_button)
        title = QLabel("Складской учет")
        title.setProperty('role', 'title')
        top_row.addWidget(title)
        top_row.addStretch()
        layout.insertLayout(0, top_row)

        # --- Новая панель фильтров ---
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(16)
        filter_layout.setContentsMargins(0, 0, 0, 0)
        # Поиск
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Поиск по названию...")
        self.search_input.setFixedWidth(260)
        self.search_input.textChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.search_input)
        # Сортировка
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["По названию", "По количеству", "По розничной цене", "По закупочной цене"])
        self.sort_combo.setFixedWidth(200)
        self.sort_combo.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Сортировка:"))
        filter_layout.addWidget(self.sort_combo)
        # Категории
        self.category_combo = QComboBox()
        self.category_combo.setFixedWidth(200)
        self.category_combo.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Категория:"))
        filter_layout.addWidget(self.category_combo)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # --- Информационный блок ---
        self.info_label = QLabel()
        self.info_label.setStyleSheet("font-size: 16px; color: #43e97b; margin-bottom: 8px;")
        layout.addWidget(self.info_label)

        # Кнопки управления (две строки)
        buttons_row1 = QHBoxLayout()
        buttons_row1.setSpacing(10)
        buttons_row1.setContentsMargins(0, 0, 0, 0)
        self.add_button = QPushButton("➕ Добавить товар")
        self.add_button.setMinimumWidth(220)
        self.add_button.setMaximumWidth(300)
        self.add_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.add_button.clicked.connect(self.add_product)
        buttons_row1.addWidget(self.add_button)

        # Новая кнопка: Добавить товар по штрихкоду
        self.add_by_barcode_button = QPushButton("🏷️ Добавить товар по штрихкоду")
        self.add_by_barcode_button.setMinimumWidth(260)
        self.add_by_barcode_button.setMaximumWidth(340)
        self.add_by_barcode_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.add_by_barcode_button.clicked.connect(self.add_product_by_barcode)
        buttons_row1.addWidget(self.add_by_barcode_button)

        self.add_supplier_button = QPushButton("🚚 Добавить поставщика")
        self.add_supplier_button.setMinimumWidth(220)
        self.add_supplier_button.setMaximumWidth(300)
        self.add_supplier_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.add_supplier_button.clicked.connect(self.add_supplier)
        buttons_row1.addWidget(self.add_supplier_button)
        self.edit_button = QPushButton("✏️ Редактировать")
        self.edit_button.setMinimumWidth(220)
        self.edit_button.setMaximumWidth(300)
        self.edit_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.edit_button.clicked.connect(self.edit_product)
        buttons_row1.addWidget(self.edit_button)
        self.delete_button = QPushButton("🗑️ Удалить")
        self.delete_button.setMinimumWidth(220)
        self.delete_button.setMaximumWidth(300)
        self.delete_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.delete_button.clicked.connect(self.delete_product)
        buttons_row1.addWidget(self.delete_button)
        self.add_category_button = QPushButton("➕ Категория")
        self.add_category_button.setMinimumWidth(220)
        self.add_category_button.setMaximumWidth(300)
        self.add_category_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.add_category_button.clicked.connect(self.show_add_category_dialog)
        buttons_row1.addWidget(self.add_category_button)
        self.delete_category_button = QPushButton("🗑️ Категория")
        self.delete_category_button.setMinimumWidth(220)
        self.delete_category_button.setMaximumWidth(300)
        self.delete_category_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.delete_category_button.clicked.connect(self.show_delete_category_dialog)
        buttons_row1.addWidget(self.delete_category_button)
        self.history_button = QPushButton("📦 История движения")
        self.history_button.setMinimumWidth(220)
        self.history_button.setMaximumWidth(300)
        self.history_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.history_button.clicked.connect(self.show_movement_history)
        buttons_row1.addWidget(self.history_button)
        buttons_row1.addStretch()
        layout.addLayout(buttons_row1)
        
        buttons_row2 = QHBoxLayout()
        buttons_row2.setSpacing(10)
        buttons_row2.setContentsMargins(0, 0, 0, 0)
        self.import_button = QPushButton("⬆️ Импорт из Excel")
        self.import_button.setMinimumWidth(220)
        self.import_button.setMaximumWidth(260)
        self.import_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.import_button.clicked.connect(self.import_products_dialog)
        buttons_row2.addWidget(self.import_button)
        self.supplier_orders_button = QPushButton("📋 Заказы у поставщика")
        self.supplier_orders_button.setMinimumWidth(220)
        self.supplier_orders_button.setMaximumWidth(260)
        self.supplier_orders_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.supplier_orders_button.clicked.connect(self.show_supplier_orders)
        buttons_row2.addWidget(self.supplier_orders_button)
        self.report_button = QPushButton("📊 Составить отчёт для заказа")
        self.report_button.setMinimumWidth(220)
        self.report_button.setMaximumWidth(260)
        self.report_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.report_button.clicked.connect(self.create_low_stock_report)
        buttons_row2.addWidget(self.report_button)
        self.report_by_qty_button = QPushButton("📑 Отчёт по количеству товаров")
        self.report_by_qty_button.setMinimumWidth(220)
        self.report_by_qty_button.setMaximumWidth(260)
        self.report_by_qty_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.report_by_qty_button.clicked.connect(self.show_quantity_report_dialog)
        buttons_row2.addWidget(self.report_by_qty_button)
        self.revenue_report_button = QPushButton("📈 Отчёт по выручке")
        self.revenue_report_button.setMinimumWidth(220)
        self.revenue_report_button.setMaximumWidth(260)
        self.revenue_report_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.revenue_report_button.clicked.connect(self.show_revenue_report_dialog)
        buttons_row2.addWidget(self.revenue_report_button)
        buttons_row2.addStretch()
        layout.addLayout(buttons_row2)

        # Только таблица товаров
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(7)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Название", "Штрихкод", "Закупочная цена", "Розничная цена", "Количество", "Категория"
        ])
        header = self.products_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)           # Название
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents) # Штрихкод
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents) # Закупочная цена
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents) # Розничная цена
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents) # Количество
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents) # Категория
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.products_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.products_table.setAlternatingRowColors(True)
        layout.addWidget(self.products_table)

    def load_products(self):
        try:
            self.db.cursor.execute("""
                SELECT id, name, barcode, purchase_price, retail_price, quantity, category 
                FROM products 
                ORDER BY name
            """)
            products = self.db.cursor.fetchall()
            self.all_products = products  # Сохраняем все товары для фильтрации
            # Загружаем категории для фильтра
            self.db.cursor.execute("SELECT name FROM categories ORDER BY name")
            categories = [row[0] for row in self.db.cursor.fetchall()]
            self.category_combo.blockSignals(True)
            self.category_combo.clear()
            self.category_combo.addItem("Все категории")
            for cat in categories:
                self.category_combo.addItem(cat)
            self.category_combo.blockSignals(False)
            self.apply_filters()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить товары: {str(e)}")

    def apply_filters(self):
        # Фильтрация
        search_text = self.search_input.text().lower()
        selected_category = self.category_combo.currentText()
        filtered = []
        for product in getattr(self, 'all_products', []):
            name = str(product[1]).lower()
            category = str(product[6]) if product[6] else ""
            if search_text and search_text not in name:
                continue
            if selected_category != "Все категории" and category != selected_category:
                continue
            filtered.append(product)
        # Сортировка
        sort_option = self.sort_combo.currentText()
        if sort_option == "По названию":
            filtered.sort(key=lambda x: str(x[1]).lower())
        elif sort_option == "По количеству":
            filtered.sort(key=lambda x: int(x[5]), reverse=True)
        elif sort_option == "По розничной цене":
            def safe_float(val):
                try: return float(val)
                except: return 0
            filtered.sort(key=lambda x: safe_float(x[4]), reverse=True)
        elif sort_option == "По закупочной цене":
            def safe_float(val):
                try: return float(val)
                except: return 0
            filtered.sort(key=lambda x: safe_float(x[3]), reverse=True)
        # Обновляем таблицу
        self.products_table.setRowCount(len(filtered))
        for row, product in enumerate(filtered):
            for col, value in enumerate(product):
                item = QTableWidgetItem(str(value))
                if col == 1:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                    item.setData(Qt.TextWordWrap, True)
                    item.setToolTip(str(value))
                self.products_table.setItem(row, col, item)
            # Проверка на низкий остаток
            quantity = int(product[5])
            self.db.cursor.execute("""
                SELECT min_quantity FROM category_min_quantities 
                WHERE category = %s
            """, (product[6],))
            min_quantity = self.db.cursor.fetchone()
            if min_quantity and quantity <= min_quantity[0]:
                for col in range(self.products_table.columnCount()):
                    self.products_table.item(row, col).setBackground(QColor("#ff6b6b"))
        self.products_table.resizeRowsToContents()
        # --- Новый блок: обновление информации ---
        self.update_info_block(filtered)

    def update_info_block(self, filtered):
        if not filtered:
            self.info_label.setText("Нет товаров по выбранным фильтрам.")
            return
        total_products = len(filtered)
        total_qty = sum(int(row[5]) for row in filtered)
        prices = [float(row[4]) for row in filtered if row[4] not in (None, '', 'None')]
        if prices:
            avg_price = sum(prices) / len(prices)
            min_price = min(prices)
            max_price = max(prices)
            price_info = f"Средняя цена: {avg_price:.2f} ₽ | Мин: {min_price:.2f} ₽ | Макс: {max_price:.2f} ₽"
        else:
            price_info = "Цены не указаны"
        self.info_label.setText(
            f"Товаров: <b>{total_products}</b> | Всего на складе: <b>{total_qty}</b> | {price_info}"
        )

    def add_product(self):
        dialog = AddProductDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            product_data = dialog.get_product_data()
            try:
                self.db.cursor.execute("""
                    INSERT INTO products (name, price, quantity, barcode, category, purchase_price, retail_price)
                    VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
                """, (
                    product_data['name'],
                    product_data['price'],
                    product_data['quantity'],
                    product_data['barcode'],
                    product_data['category'],
                    product_data['purchase_price'],
                    product_data['retail_price']
                ))
                product_id = self.db.cursor.fetchone()[0]
                # Записываем движение товара
                self.db.add_product_movement(
                    product_id=product_id,
                    movement_type='IN',
                    quantity=int(product_data['quantity']),
                    username=self.username,
                    comment=f'Первоначальное поступление товара'
                )
                self.db.connection.commit()
                self.load_products()
                QMessageBox.information(self, "Успех", "Товар успешно добавлен")
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить товар: {str(e)}")

    def edit_product(self):
        selected_rows = self.products_table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите товар для редактирования")
            return
        row = selected_rows[0].row()
        product_id = int(self.products_table.item(row, 0).text())
        old_quantity = int(self.products_table.item(row, 5).text())
        dialog = AddProductDialog(self.db, self)
        dialog.name_input.setText(self.products_table.item(row, 1).text())
        dialog.purchase_price_input.setText(self.products_table.item(row, 3).text())
        dialog.retail_price_input.setText(self.products_table.item(row, 4).text())
        dialog.quantity_input.setValue(old_quantity)
        dialog.category_combo.setCurrentText(self.products_table.item(row, 6).text())
        if dialog.exec_() == QDialog.Accepted:
            product_data = dialog.get_product_data()
            try:
                new_quantity = int(product_data['quantity'])
                quantity_diff = new_quantity - old_quantity
                self.db.cursor.execute("""
                    UPDATE products 
                    SET name = %s, purchase_price = %s, retail_price = %s, quantity = %s, barcode = %s, category = %s
                    WHERE id = %s
                """, (
                    product_data['name'],
                    product_data['purchase_price'],
                    product_data['retail_price'],
                    product_data['quantity'],
                    product_data['barcode'],
                    product_data['category'],
                    product_id
                ))
                if quantity_diff != 0:
                    self.db.add_product_movement(
                        product_id=product_id,
                        movement_type='IN' if quantity_diff > 0 else 'OUT',
                        quantity=abs(quantity_diff),
                        username=self.username,
                        comment=f'Ручное изменение количества с {old_quantity} на {new_quantity}'
                    )
                self.db.connection.commit()
                self.load_products()
                QMessageBox.information(self, "Успех", "Товар успешно обновлен")
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось обновить товар: {str(e)}")

    def delete_product(self):
        selected_rows = self.products_table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "Предупреждение", "Выберите товар для удаления")
            return
        
        row = selected_rows[0].row()
        product_id = int(self.products_table.item(row, 0).text())
        product_name = self.products_table.item(row, 1).text()
        current_quantity = int(self.products_table.item(row, 5).text())
        
        reply = QMessageBox.question(
            self, 
            "Подтверждение", 
            f"Вы уверены, что хотите удалить товар '{product_name}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # Записываем списание всего количества
                if current_quantity > 0:
                    self.db.add_product_movement(
                        product_id=product_id,
                        movement_type='OUT',
                        quantity=current_quantity,
                        username=self.username,
                        comment='Списание при удалении товара'
                    )
                
                self.db.cursor.execute("DELETE FROM products WHERE id = %s", (product_id,))
                self.db.connection.commit()
                self.load_products()
                QMessageBox.information(self, "Успех", "Товар успешно удален")
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось удалить товар: {str(e)}")

    def show_automation_settings(self):
        dialog = AutomationSettingsDialog(self.automation, self)
        dialog.exec_()

    def on_low_stock(self, product_name, quantity):
        if self.automation.settings['notify_on_low']:
            QMessageBox.warning(
                self,
                "Низкий остаток",
                f"Товар '{product_name}' имеет низкий остаток: {quantity} шт."
            )

    def on_order_needed(self, product_name, order_quantity):
        if self.automation.settings['auto_order']:
            reply = QMessageBox.question(
                self,
                "Формирование заказа",
                f"Сформировать заказ на товар '{product_name}' в количестве {order_quantity} шт.?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                # Здесь можно добавить логику формирования заказа
                QMessageBox.information(
                    self,
                    "Заказ сформирован",
                    f"Заказ на товар '{product_name}' в количестве {order_quantity} шт. успешно сформирован."
                )

    def show_price_list_dialog(self):
        dialog = PriceListDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_products()  # Обновляем список товаров после добавления новых

    def show_supplier_orders(self):
        # Добавляю кнопку возврата к складу
        if not hasattr(self, 'back_to_stock_btn'):
            self.back_to_stock_btn = QPushButton("Склад")
            self.back_to_stock_btn.setStyleSheet("background-color: #3fa996; color: #23243a; font-weight: bold; padding: 4px 18px; border-radius: 8px; font-size: 15px; min-width: 80px; min-height: 28px; max-width: 120px; max-height: 32px;")
            self.back_to_stock_btn.clicked.connect(self.show_products_table)
            self.layout().insertWidget(0, self.back_to_stock_btn)
        else:
            self.back_to_stock_btn.show()
        # Скрываем кнопки управления товарами
        self.add_button.hide()
        self.edit_button.hide()
        self.delete_button.hide()
        self.products_table.setColumnCount(5)
        self.products_table.setHorizontalHeaderLabels([
            "ID заказа", "Поставщик", "Количество", "Сумма заказа", "Статус"
        ])
        # --- Stretch только для колонки 'Поставщик' ---
        header = self.products_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID заказа
        header.setSectionResizeMode(1, QHeaderView.Stretch)           # Поставщик
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Количество
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Сумма заказа
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Статус
        self.db.cursor.execute("""
            SELECT po.id, s.name, 
                   COALESCE(SUM(poi.quantity), 0) as total_qty, 
                   COALESCE(SUM(poi.quantity * poi.price), 0) as total_sum,
                   po.status
            FROM pending_orders po
            LEFT JOIN suppliers s ON po.supplier = s.id
            LEFT JOIN pending_order_items poi ON poi.order_id = po.id
            GROUP BY po.id, s.name, po.status
            ORDER BY po.id DESC
        """)
        orders = self.db.cursor.fetchall()
        self.products_table.setRowCount(len(orders))
        for row, order in enumerate(orders):
            for col, value in enumerate(order):
                item = QTableWidgetItem(str(value))
                # --- Новое: перенос строк для названия поставщика (если нужно) ---
                if col == 1:  # "Поставщик"
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    item.setData(Qt.TextWordWrap, True)
                    item.setToolTip(str(value))
                self.products_table.setItem(row, col, item)
        # --- Новое: подгоняем высоту строк под содержимое ---
        self.products_table.resizeRowsToContents()
        # Подключаем обработчик правой кнопки мыши
        self.products_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.products_table.customContextMenuRequested.connect(self.open_order_context_menu)

        # Добавляем кнопку "Отметить поступление" под таблицей
        if not hasattr(self, 'mark_received_btn'):
            self.mark_received_btn = QPushButton("Отметить поступление")
            self.mark_received_btn.setStyleSheet("background-color: #43e97b; color: #23243a; font-weight: bold; padding: 10px 24px; border-radius: 8px;")
            self.mark_received_btn.clicked.connect(self.mark_order_received)
            self.layout().addWidget(self.mark_received_btn)
        else:
            self.mark_received_btn.show()

    def open_order_context_menu(self, position):
        indexes = self.products_table.selectedIndexes()
        if not indexes:
            return
        row = self.products_table.rowAt(position.y())
        if row < 0:
            return
        menu = QMenu()
        info_action = QAction("Просмотреть информацию", self)
        menu.addAction(info_action)
        action = menu.exec_(self.products_table.viewport().mapToGlobal(position))
        if action == info_action:
            order_id = self.products_table.item(row, 0).text()
            self.show_order_details(order_id)

    def show_order_details(self, order_id):
        # Получаем товары по заказу
        self.db.cursor.execute("""
            SELECT name, price, quantity
            FROM pending_order_items
            WHERE order_id = %s
        """, (order_id,))
        items = self.db.cursor.fetchall()
        # Формируем текст для отображения
        details = ""
        for item in items:
            details += f"Имя товара: {item[0]}\nЦена: {item[1]}\nКоличество: {item[2]}\n---\n"
        if not details:
            details = "Нет товаров в этом заказе."
        QMessageBox.information(self, "Детали заказа", details)

    def show_products_table(self):
        # Возвращаем таблицу товаров
        self.products_table.setColumnCount(7)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Название", "Штрихкод", "Закупочная цена", "Розничная цена", "Количество", "Категория"
        ])
        # Показываем кнопки управления товарами
        self.add_button.show()
        self.edit_button.show()
        self.delete_button.show()
        self.load_products()
        # Скрываем кнопку возврата, если она есть
        if hasattr(self, 'back_to_stock_btn'):
            self.back_to_stock_btn.hide()
        # Скрываем кнопку "Отметить поступление", если она есть
        if hasattr(self, 'mark_received_btn'):
            self.mark_received_btn.hide()

    def add_supplier(self):
        dialog = AddSupplierDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            supplier_data = dialog.get_supplier_data()
            if not supplier_data['name'].strip():
                QMessageBox.warning(self, "Ошибка", "Название поставщика не может быть пустым!")
                return
            if self.db.add_supplier(supplier_data):
                QMessageBox.information(self, "Поставщик добавлен", f"Поставщик '{supplier_data['name']}' успешно добавлен!")
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось добавить поставщика. Проверьте корректность данных.")

    def import_products_dialog(self):
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл для импорта", "", "Excel/CSV Files (*.xlsx *.xls *.csv)")
        if not file_path:
            return
        # Загружаем файл в DataFrame
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {e}")
            return
        # Передаём весь DataFrame в ColumnMappingDialog
        dialog = ColumnMappingDialog(df, parent=self, excel_file=file_path)
        if dialog.exec_() == QDialog.Accepted:
            mapping = dialog.get_mapping()
            col_map = dict(zip(dialog.df.columns, mapping))
            if "Название" not in mapping or "Цена" not in mapping:
                QMessageBox.warning(self, "Ошибка", "Не выбраны все обязательные поля (Название, Цена)!")
                return
            # --- Новый блок: оформление заказа у поставщика ---
            # 1. Диалог выбора поставщика
            supplier_dialog = SupplierSelectDialog(self.db, self)
            if supplier_dialog.exec_() != QDialog.Accepted:
                return
            supplier_id, supplier_name = supplier_dialog.get_selected_supplier()
            # 2. Формируем список товаров для заказа
            name_col = None
            price_col = None
            qty_col = None
            for col, mapped in col_map.items():
                if mapped == "Название":
                    name_col = col
                elif mapped == "Цена":
                    price_col = col
                elif mapped == "Количество":
                    qty_col = col
            if not name_col or not price_col:
                QMessageBox.warning(self, "Ошибка", "Не выбраны все обязательные поля (Название, Цена)!")
                return
            try:
                # 3.1. Считаем сумму заказа и общее количество
                total_sum = 0
                total_qty = 0
                print("\n=== ОТЛАДКА ИМПОРТА ЗАКАЗА ===")
                print("Столбцы DataFrame:", list(dialog.df.columns))
                print("Сопоставление:", col_map)
                # Проверяем, что выбранные столбцы существуют
                if name_col not in dialog.df.columns or price_col not in dialog.df.columns or (qty_col and qty_col not in dialog.df.columns):
                    print(f"[ОШИБКА] Не найден столбец: name_col={name_col}, price_col={price_col}, qty_col={qty_col}")
                    QMessageBox.critical(self, "Ошибка", "Проверьте сопоставление колонок! Возможно, выбран пустой или несуществующий столбец.")
                    return
                for idx, row in dialog.df.iterrows():
                    print(f"Строка {idx}: {row.to_dict()}")
                    try:
                        name = str(row[name_col])
                    except Exception:
                        print(f"[ОШИБКА] Не найден столбец: {name_col} в строке {row.to_dict()}")
                        QMessageBox.critical(self, "Ошибка", f"Проверьте сопоставление колонок! Не найден столбец: {name_col}")
                        return
                    try:
                        price = float(row[price_col])
                    except Exception:
                        price = 0.0
                    qty = 1
                    if qty_col:
                        try:
                            qty = int(float(row[qty_col]))
                        except Exception:
                            qty = 1
                    if price > 0 and qty > 0:
                        total_sum += price * qty
                        total_qty += qty
                print(f"Итого: total_sum={total_sum}, total_qty={total_qty}")
                # 3.2. Добавляем заказ (шапку)
                self.db.cursor.execute("""
                    INSERT INTO pending_orders (name, supplier, price, quantity, order_date, status)
                    VALUES (%s, %s, %s, %s, NOW(), %s) RETURNING id
                """, (supplier_name, supplier_id, total_sum, total_qty, 'В процессе'))
                order_id = self.db.cursor.fetchone()[0]
                # 3.3. Добавляем товары (позиции)
                for idx, row in dialog.df.iterrows():
                    name = str(row[name_col])
                    try:
                        price = float(row[price_col])
                    except Exception:
                        price = 0.0
                    qty = 1
                    if qty_col:
                        try:
                            qty = int(float(row[qty_col]))
                        except Exception:
                            qty = 1
                    if not name or price <= 0 or qty <= 0:
                        continue
                    self.db.cursor.execute("""
                        INSERT INTO pending_order_items (order_id, name, price, quantity, category)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (
                        order_id,
                        name,
                        price,
                        qty,
                        row[col_map.get("Категория")] if "Категория" in col_map and col_map["Категория"] in row else "Без категории"
                    ))
                self.db.connection.commit()
                QMessageBox.information(self, "Успех", f"Заказ успешно создан для поставщика: {supplier_name}")
                self.show_supplier_orders()
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось создать заказ: {e}") 

    def mark_order_received(self):
        selected = self.products_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Внимание", "Выберите заказ для отметки поступления!")
            return
        row = selected[0].row()
        order_id = self.products_table.item(row, 0).text()
        try:
            # Получаем все товары по заказу
            self.db.cursor.execute("SELECT name, price, quantity, category FROM pending_order_items WHERE order_id = %s", (order_id,))
            items = self.db.cursor.fetchall()
            for name, price, qty, category in items:
                # Проверяем наличие товара по штрихкоду или названию
                self.db.cursor.execute("SELECT id, quantity FROM products WHERE name = %s OR barcode = %s", (name, name))
                product = self.db.cursor.fetchone()
                if product:
                    # Товар есть, увеличиваем количество
                    prod_id, old_qty = product
                    new_qty = int(old_qty) + int(qty)
                    self.db.cursor.execute("UPDATE products SET quantity = %s WHERE id = %s", (new_qty, prod_id))
                    
                    # Записываем движение товара
                    self.db.add_product_movement(
                        product_id=prod_id,
                        movement_type='IN',
                        quantity=int(qty),
                        username=self.username,
                        comment=f'Поступление по заказу #{order_id}'
                    )
                else:
                    # Товара нет — открываем AddProductDialog с автозаполнением
                    dialog = AddProductDialog(self.db, self)
                    dialog.name_input.setText(str(name))
                    dialog.price_input.setText(str(price))
                    dialog.quantity_input.setValue(int(qty))
                    dialog.category_combo.setCurrentText(str(category))
                    # barcode оставляем пустым, пользователь может ввести
                    if dialog.exec_() == QDialog.Accepted:
                        product_data = dialog.get_product_data()
                        self.db.cursor.execute(
                            """INSERT INTO products (name, price, quantity, barcode, category) 
                               VALUES (%s, %s, %s, %s, %s) RETURNING id""",
                            (
                                product_data['name'],
                                product_data['price'],
                                product_data['quantity'],
                                product_data['barcode'],
                                product_data['category']
                            )
                        )
                        product_id = self.db.cursor.fetchone()[0]
                        
                        # Записываем движение товара
                        self.db.add_product_movement(
                            product_id=product_id,
                            movement_type='IN',
                            quantity=int(product_data['quantity']),
                            username=self.username,
                            comment=f'Первое поступление по заказу #{order_id}'
                        )
            # После обработки всех товаров — отмечаем заказ как поступивший
            self.db.cursor.execute("UPDATE pending_orders SET status = %s WHERE id = %s", ("Поступил", order_id))
            self.db.connection.commit()
            self.load_products()
            QMessageBox.information(self, "Успех", "Заказ успешно отмечен как поступивший")
        except Exception as e:
            self.db.connection.rollback()
            QMessageBox.critical(self, "Ошибка", f"Не удалось обработать поступление заказа: {str(e)}")

    def create_low_stock_report(self):
        import pandas as pd
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        # Получаем товары с низким остатком
        low_stock = self.db.get_low_stock_products()
        if not low_stock:
            QMessageBox.information(self, "Отчёт", "Нет товаров с низким остатком!")
            return
        df = pd.DataFrame(low_stock)
        # Удаляем столбец 'min_quantity'/'Минимальное количество' по всем вариантам
        for col in list(df.columns):
            if str(col).strip().lower() in ["min_quantity", "минимальное количество"]:
                df.drop(columns=[col], inplace=True)
        # Диалог выбора файла
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт", "отчёт_по_заказу.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            df.rename(columns={
                'name': 'Товар',
                'category': 'Категория',
                'quantity': 'Текущее количество',
            }, inplace=True)
            df.to_excel(file_path, index=False)
            # Стилизация: увеличиваем шрифт и автоширина для всех столбцов
            wb = load_workbook(file_path)
            ws = wb.active
            font = Font(size=14)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                ws.row_dimensions[row[0].row].height = 28
            # Автоширина для всех столбцов
            for i, col in enumerate(ws.columns):
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                if i == 0:  # Первая колонка — 'Товар'
                    ws.column_dimensions[col[0].column_letter].width = max_len + 20  # Без ограничения, с запасом
                else:
                    ws.column_dimensions[col[0].column_letter].width = max(18, min(max_len + 6, 80))
            wb.save(file_path)
            QMessageBox.information(self, "Отчёт", f"Отчёт успешно сохранён: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить отчёт: {e}") 

    def show_movement_history(self):
        selected_rows = self.products_table.selectedItems()
        product_id = None
        if selected_rows:
            row = selected_rows[0].row()
            product_id = int(self.products_table.item(row, 0).text())
        
        from app_code.dialogs import ProductMovementHistoryDialog
        dialog = ProductMovementHistoryDialog(self.db, product_id, self)
        dialog.exec_()

    def show_quantity_report_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Выберите категории для отчёта")
        dialog.resize(500, 420)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #23243a;
                border-radius: 18px;
                border: 2px solid #43e97b;
            }
            QListWidget {
                background: #292a3a;
                color: #f3f3f3;
                border-radius: 10px;
                border: 1.5px solid #43e97b;
                font-size: 18px;
                padding: 8px 14px;
                margin-bottom: 8px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #43e97b, stop:1 #2fd072);
                color: #23243a;
                font-weight: bold;
                border-radius: 10px;
                font-size: 18px;
                padding: 10px 0;
                margin: 8px 0 0 0;
            }
            QPushButton:hover {
                background: #43e97b;
                color: #23243a;
            }
            QLabel {
                font-size: 17px;
            }
        """)
        vbox = QVBoxLayout(dialog)
        label = QLabel("Выберите одну, несколько или все категории:")
        vbox.addWidget(label)
        category_list = QListWidget()
        category_list.setSelectionMode(QListWidget.MultiSelection)
        self.db.cursor.execute("SELECT name FROM categories ORDER BY name")
        categories = [row[0] for row in self.db.cursor.fetchall()]
        for cat in categories:
            item = QListWidgetItem(cat)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            category_list.addItem(item)
        vbox.addWidget(category_list)
        btns = QHBoxLayout()
        ok_btn = QPushButton("Сформировать отчёт")
        cancel_btn = QPushButton("Отмена")
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        vbox.addLayout(btns)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        if dialog.exec_() == QDialog.Accepted:
            selected = [category_list.item(i).text() for i in range(category_list.count()) if category_list.item(i).checkState() == Qt.Checked]
            if not selected:
                selected = categories  # Если ничего не выбрано — все категории
            self.create_quantity_report(selected)

    def create_quantity_report(self, selected_categories):
        import pandas as pd
        from PyQt5.QtWidgets import QFileDialog
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        # Получаем товары выбранных категорий
        placeholders = ','.join(['%s'] * len(selected_categories))
        self.db.cursor.execute(f"""
            SELECT name, category, quantity FROM products
            WHERE category IN ({placeholders})
            ORDER BY category, name
        """, tuple(selected_categories))
        products = self.db.cursor.fetchall()
        if not products:
            QMessageBox.information(self, "Отчёт", "Нет товаров для выбранных категорий!")
            return
        df = pd.DataFrame(products, columns=["Название", "Категория", "Количество"])
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт", "отчёт_по_количеству.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            df.to_excel(file_path, index=False)
            # Оформление через openpyxl
            wb = load_workbook(file_path)
            ws = wb.active
            # Заголовки жирные и крупные
            header_font = Font(size=14, bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Границы для всех ячеек
            thin = Side(border_style="thin", color="888888")
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Авто-перенос текста и ширина для всех столбцов
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                ws.column_dimensions[col[0].column_letter].width = max(18, min(max_len + 8, 60))
            # Высота строк под содержимое
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 32
            # Автоширина для остальных столбцов
            for col in ws.columns:
                if col[0].column_letter == 'A':
                    continue
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 4, 40))
            # Добавляем строку "Итого" внизу
            last_row = ws.max_row + 1
            ws[f'A{last_row}'] = "Итого:"
            ws[f'A{last_row}'].font = Font(size=14, bold=True)
            ws[f'E{last_row}'] = df["Количество"].sum()
            ws[f'E{last_row}'].font = Font(size=14, bold=True)
            ws[f'E{last_row}'].number_format = '#,##0.00'
            wb.save(file_path)
            QMessageBox.information(self, "Отчёт по выручке", f"Отчёт успешно сохранён!\nОбщая выручка: {df['Количество'].sum():.2f} ₽")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить отчёт: {e}") 

    def show_add_category_dialog(self):
        text, ok = QInputDialog.getText(self, "Добавить категорию", "Введите название новой категории:")
        if ok and text.strip():
            try:
                self.db.cursor.execute("INSERT INTO categories (name) VALUES (%s)", (text.strip(),))
                self.db.connection.commit()
                QMessageBox.information(self, "Успех", f"Категория '{text.strip()}' успешно добавлена!")
                self.load_products()
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить категорию: {e}")

    def show_delete_category_dialog(self):
        self.db.cursor.execute("SELECT name FROM categories ORDER BY name")
        categories = [row[0] for row in self.db.cursor.fetchall()]
        if not categories:
            QMessageBox.information(self, "Удаление категории", "Нет категорий для удаления.")
            return
        item, ok = QInputDialog.getItem(self, "Удалить категорию", "Выберите категорию для удаления:", categories, 0, False)
        if ok and item:
            reply = QMessageBox.question(self, "Подтверждение", f"Удалить категорию '{item}'? Все товары с этой категорией останутся без категории.", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    self.db.cursor.execute("UPDATE products SET category=NULL WHERE category=%s", (item,))
                    self.db.cursor.execute("DELETE FROM categories WHERE name=%s", (item,))
                    self.db.connection.commit()
                    QMessageBox.information(self, "Успех", f"Категория '{item}' удалена.")
                    self.load_products()
                except Exception as e:
                    self.db.connection.rollback()
                    QMessageBox.critical(self, "Ошибка", f"Не удалось удалить категорию: {e}") 

    def show_revenue_report_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Отчёт по выручке")
        dialog.setModal(True)
        dialog.resize(340, 180)
        dialog.setStyleSheet("""
            QDialog { background: #23243a; border-radius: 14px; border: 2px solid #3fa996; }
            QLabel { color: #3fa996; font-size: 18px; font-weight: bold; }
            QComboBox { background: #292a3a; color: #f3f3f3; border-radius: 8px; border: 1.5px solid #3fa996; font-size: 17px; padding: 8px 14px; }
            QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3fa996, stop:1 #2d7a6a); color: #f3f3f3; font-weight: bold; border-radius: 10px; font-size: 18px; padding: 10px 0; margin-top: 12px; }
            QPushButton:hover { background: #4fd1b3; color: #23243a; }
        """)
        vbox = QVBoxLayout(dialog)
        vbox.addWidget(QLabel("Выберите период:"))
        self.period_combo = QComboBox(dialog)
        self.period_combo.addItems(["День", "Неделя", "Месяц"])
        vbox.addWidget(self.period_combo)
        btn = QPushButton("Сформировать")
        btn.clicked.connect(lambda: (dialog.accept(), self.generate_revenue_report(self.period_combo.currentText())))
        vbox.addWidget(btn)
        dialog.exec_()

    def generate_revenue_report(self, period):
        # Здесь будет логика формирования отчёта по выручке за выбранный период
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        import pandas as pd
        import datetime
        # Определяем границы периода
        now = datetime.datetime.now()
        if period == "День":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0).date()
        elif period == "Неделя":
            start = (now - datetime.timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0).date()
        elif period == "Месяц":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).date()
        else:
            QMessageBox.warning(self, "Ошибка", "Неизвестный период!")
            return
        # Получаем продажи за период с закупочной и розничной ценой
        self.db.cursor.execute("""
            SELECT sh.sale_date, sh.product_name, sh.quantity, sh.sale_price, p.purchase_price, p.retail_price
            FROM sales_history sh
            LEFT JOIN products p ON sh.product_name = p.name
            WHERE sh.sale_date::date >= %s
            ORDER BY sh.sale_date
        """, (str(start),))
        sales = self.db.cursor.fetchall()
        if not sales:
            QMessageBox.information(self, "Отчёт по выручке", "Нет продаж за выбранный период.")
            return
        df = pd.DataFrame(sales, columns=["Дата", "Товар", "Количество", "Цена продажи", "Закупочная цена", "Розничная цена"])
        # Выручка = (Цена продажи - Закупочная цена) * Количество
        df["Выручка"] = (pd.to_numeric(df["Цена продажи"], errors='coerce').fillna(0) - pd.to_numeric(df["Закупочная цена"], errors='coerce').fillna(0)) * pd.to_numeric(df["Количество"], errors='coerce').fillna(0)
        total_revenue = df["Выручка"].sum()
        # Сохраняем отчёт
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт", f"выручка_{period.lower()}.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            df.to_excel(file_path, index=False)
            # Оформление через openpyxl
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            wb = load_workbook(file_path)
            ws = wb.active
            # Заголовки жирные и крупные
            header_font = Font(size=14, bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Границы для всех ячеек
            thin = Side(border_style="thin", color="888888")
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Авто-перенос текста и ширина для всех столбцов
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                ws.column_dimensions[col[0].column_letter].width = max(18, min(max_len + 8, 60))
            # Высота строк под содержимое
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 32
            # Добавляем строку "Итого" внизу
            last_row = ws.max_row + 1
            ws[f'A{last_row}'] = "Итого:"
            ws[f'A{last_row}'].font = Font(size=14, bold=True)
            ws[f'G{last_row}'] = total_revenue
            ws[f'G{last_row}'].font = Font(size=14, bold=True)
            ws[f'G{last_row}'].number_format = '#,##0.00'
            wb.save(file_path)
            QMessageBox.information(self, "Отчёт по выручке", f"Отчёт успешно сохранён!\nОбщая выручка: {total_revenue:.2f} ₽")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить отчёт: {e}") 

    def add_product_by_barcode(self):
        dialog = AddProductDialog(self.db, self)
        dialog.barcode_input.setFocus()
        if dialog.exec_() == QDialog.Accepted:
            product_data = dialog.get_product_data()
            try:
                self.db.cursor.execute("""
                    INSERT INTO products (name, price, quantity, barcode, category, purchase_price, retail_price)
                    VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
                """, (
                    product_data['name'],
                    product_data['price'],
                    product_data['quantity'],
                    product_data['barcode'],
                    product_data['category'],
                    product_data['purchase_price'],
                    product_data['retail_price']
                ))
                product_id = self.db.cursor.fetchone()[0]
                self.db.add_product_movement(
                    product_id=product_id,
                    movement_type='IN',
                    quantity=int(product_data['quantity']),
                    username=self.username,
                    comment=f'Поступление по штрихкоду'
                )
                self.db.connection.commit()
                self.load_products()
                QMessageBox.information(self, "Успех", "Товар успешно добавлен по штрихкоду")
            except Exception as e:
                self.db.connection.rollback()
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить товар: {str(e)}") 